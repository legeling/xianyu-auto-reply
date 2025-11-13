"""业务管理相关的大部分路由（通知、系统、关键词、商品、日志、管理员等）。"""

from __future__ import annotations

import glob
import io
import json
import os
import shutil
import sqlite3
import time
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from loguru import logger
from pydantic import BaseModel

from app.api.dependencies import (
    get_current_user,
    get_current_user_optional,
    log_with_user,
    require_auth,
    require_admin,
)
from app.repositories.db_manager import db_manager
from app.services import cookie_manager
from app.services.ai_reply import ai_reply_engine
from app.services.file_log_collector import get_file_log_collector
from app.utils.image_utils import image_manager

router = APIRouter()

# ------------------------- 通用数据模型 -------------------------


class NotificationChannelIn(BaseModel):
    name: str
    type: str = "qq"
    config: str


class NotificationChannelUpdate(BaseModel):
    name: str
    config: str
    enabled: bool = True


class MessageNotificationIn(BaseModel):
    channel_id: int
    enabled: bool = True


class SystemSettingIn(BaseModel):
    value: str
    description: Optional[str] = ""


# ------------------------- 通知渠道管理接口 -------------------------

# ------------------------- 通知渠道管理接口 -------------------------

@router.get('/notification-channels')
def get_notification_channels(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取所有通知渠道"""
    from app.repositories.db_manager import db_manager
    try:
        user_id = current_user['user_id']
        return db_manager.get_notification_channels(user_id)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post('/notification-channels')
def create_notification_channel(channel_data: NotificationChannelIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建通知渠道"""
    from app.repositories.db_manager import db_manager
    try:
        user_id = current_user['user_id']
        channel_id = db_manager.create_notification_channel(
            channel_data.name,
            channel_data.type,
            channel_data.config,
            user_id
        )
        return {'msg': 'notification channel created', 'id': channel_id}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get('/notification-channels/{channel_id}')
def get_notification_channel(channel_id: int, _: None = Depends(require_auth)):
    """获取指定通知渠道"""
    from app.repositories.db_manager import db_manager
    try:
        channel = db_manager.get_notification_channel(channel_id)
        if not channel:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
        return channel
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put('/notification-channels/{channel_id}')
def update_notification_channel(channel_id: int, channel_data: NotificationChannelUpdate, _: None = Depends(require_auth)):
    """更新通知渠道"""
    from app.repositories.db_manager import db_manager
    try:
        success = db_manager.update_notification_channel(
            channel_id,
            channel_data.name,
            channel_data.config,
            channel_data.enabled
        )
        if success:
            return {'msg': 'notification channel updated'}
        else:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete('/notification-channels/{channel_id}')
def delete_notification_channel(channel_id: int, _: None = Depends(require_auth)):
    """删除通知渠道"""
    from app.repositories.db_manager import db_manager
    try:
        success = db_manager.delete_notification_channel(channel_id)
        if success:
            return {'msg': 'notification channel deleted'}
        else:
            raise HTTPException(status_code=404, detail='通知渠道不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 消息通知配置接口 -------------------------

@router.get('/message-notifications')
def get_all_message_notifications(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的消息通知配置"""
    from app.repositories.db_manager import db_manager
    try:
        # 只返回当前用户的消息通知配置
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        all_notifications = db_manager.get_all_message_notifications()
        # 过滤只属于当前用户的通知配置
        user_notifications = {cid: notifications for cid, notifications in all_notifications.items() if cid in user_cookies}
        return user_notifications
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get('/message-notifications/{cid}')
def get_account_notifications(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的消息通知配置"""
    from app.repositories.db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        return db_manager.get_account_notifications(cid)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post('/message-notifications/{cid}')
def set_message_notification(cid: str, notification_data: MessageNotificationIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """设置账号的消息通知"""
    from app.repositories.db_manager import db_manager
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 检查通知渠道是否存在
        channel = db_manager.get_notification_channel(notification_data.channel_id)
        if not channel:
            raise HTTPException(status_code=404, detail='通知渠道不存在')

        success = db_manager.set_message_notification(cid, notification_data.channel_id, notification_data.enabled)
        if success:
            return {'msg': 'message notification set'}
        else:
            raise HTTPException(status_code=400, detail='设置失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete('/message-notifications/account/{cid}')
def delete_account_notifications(cid: str, _: None = Depends(require_auth)):
    """删除账号的所有消息通知配置"""
    from app.repositories.db_manager import db_manager
    try:
        success = db_manager.delete_account_notifications(cid)
        if success:
            return {'msg': 'account notifications deleted'}
        else:
            raise HTTPException(status_code=404, detail='账号通知配置不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete('/message-notifications/{notification_id}')
def delete_message_notification(notification_id: int, _: None = Depends(require_auth)):
    """删除消息通知配置"""
    from app.repositories.db_manager import db_manager
    try:
        success = db_manager.delete_message_notification(notification_id)
        if success:
            return {'msg': 'message notification deleted'}
        else:
            raise HTTPException(status_code=404, detail='通知配置不存在')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 系统设置接口 -------------------------

@router.get('/system-settings')
def get_system_settings(_: None = Depends(require_auth)):
    """获取系统设置（排除敏感信息）"""
    from app.repositories.db_manager import db_manager
    try:
        settings = db_manager.get_all_system_settings()
        # 移除敏感信息
        if 'admin_password_hash' in settings:
            del settings['admin_password_hash']
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





@router.put('/system-settings/{key}')
def update_system_setting(key: str, setting_data: SystemSettingIn, _: None = Depends(require_auth)):
    """更新系统设置"""
    from app.repositories.db_manager import db_manager
    try:
        # 禁止直接修改密码哈希
        if key == 'admin_password_hash':
            raise HTTPException(status_code=400, detail='请使用密码修改接口')

        success = db_manager.set_system_setting(key, setting_data.value, setting_data.description)
        if success:
            return {'msg': 'system setting updated'}
        else:
            raise HTTPException(status_code=400, detail='更新失败')
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 注册设置接口 -------------------------

@router.get('/registration-status')
def get_registration_status():
    """获取注册开关状态（公开接口，无需认证）"""
    from app.repositories.db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('registration_enabled')
        logger.info(f"从数据库获取的注册设置值: '{enabled_str}'")  # 调试信息

        # 如果设置不存在，默认为开启
        if enabled_str is None:
            enabled_bool = True
            message = '注册功能已开启'
        else:
            enabled_bool = enabled_str == 'true'
            message = '注册功能已开启' if enabled_bool else '注册功能已关闭'

        logger.info(f"解析后的注册状态: enabled={enabled_bool}, message='{message}'")  # 调试信息

        return {
            'enabled': enabled_bool,
            'message': message
        }
    except Exception as e:
        logger.error(f"获取注册状态失败: {e}")
        return {'enabled': True, 'message': '注册功能已开启'}  # 出错时默认开启


@router.get('/login-info-status')
def get_login_info_status():
    """获取默认登录信息显示状态（公开接口，无需认证）"""
    from app.repositories.db_manager import db_manager
    try:
        enabled_str = db_manager.get_system_setting('show_default_login_info')
        logger.debug(f"从数据库获取的登录信息显示设置值: '{enabled_str}'")

        # 如果设置不存在，默认为开启
        if enabled_str is None:
            enabled_bool = True
        else:
            enabled_bool = enabled_str == 'true'

        return {"enabled": enabled_bool}
    except Exception as e:
        logger.error(f"获取登录信息显示状态失败: {e}")
        # 出错时默认为开启
        return {"enabled": True}


class RegistrationSettingUpdate(BaseModel):
    enabled: bool


class LoginInfoSettingUpdate(BaseModel):
    enabled: bool


@router.put('/registration-settings')
def update_registration_settings(setting_data: RegistrationSettingUpdate, admin_user: Dict[str, Any] = Depends(require_admin)):
    """更新注册开关设置（仅管理员）"""
    from app.repositories.db_manager import db_manager
    try:
        enabled = setting_data.enabled
        success = db_manager.set_system_setting(
            'registration_enabled',
            'true' if enabled else 'false',
            '是否开启用户注册'
        )
        if success:
            log_with_user('info', f"更新注册设置: {'开启' if enabled else '关闭'}", admin_user)
            return {
                'success': True,
                'enabled': enabled,
                'message': f"注册功能已{'开启' if enabled else '关闭'}"
            }
        else:
            raise HTTPException(status_code=500, detail='更新注册设置失败')
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新注册设置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.put('/login-info-settings')
def update_login_info_settings(setting_data: LoginInfoSettingUpdate, admin_user: Dict[str, Any] = Depends(require_admin)):
    """更新默认登录信息显示设置（仅管理员）"""
    from app.repositories.db_manager import db_manager
    try:
        enabled = setting_data.enabled
        success = db_manager.set_system_setting(
            'show_default_login_info',
            'true' if enabled else 'false',
            '是否显示默认登录信息'
        )
        if success:
            log_with_user('info', f"更新登录信息显示设置: {'开启' if enabled else '关闭'}", admin_user)
            return {
                'success': True,
                'enabled': enabled,
                'message': f"默认登录信息显示已{'开启' if enabled else '关闭'}"
            }
        else:
            raise HTTPException(status_code=500, detail='更新登录信息显示设置失败')
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新登录信息显示设置失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))




@router.delete("/cookies/{cid}")
def remove_cookie(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        cookie_manager.manager.remove_cookie(cid)
        return {"msg": "removed"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class AutoConfirmUpdate(BaseModel):
    auto_confirm: bool


class RemarkUpdate(BaseModel):
    remark: str


class PauseDurationUpdate(BaseModel):
    pause_duration: int


@router.put("/cookies/{cid}/auto-confirm")
def update_auto_confirm(cid: str, update_data: AutoConfirmUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号的自动确认发货设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新数据库中的auto_confirm设置
        success = db_manager.update_auto_confirm(cid, update_data.auto_confirm)
        if not success:
            raise HTTPException(status_code=500, detail="更新自动确认发货设置失败")

        # 通知CookieManager更新设置（如果账号正在运行）
        if hasattr(cookie_manager.manager, 'update_auto_confirm_setting'):
            cookie_manager.manager.update_auto_confirm_setting(cid, update_data.auto_confirm)

        return {
            "msg": "success",
            "auto_confirm": update_data.auto_confirm,
            "message": f"自动确认发货已{'开启' if update_data.auto_confirm else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/cookies/{cid}/auto-confirm")
def get_auto_confirm(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号的自动确认发货设置"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取auto_confirm设置
        auto_confirm = db_manager.get_auto_confirm(cid)
        return {
            "auto_confirm": auto_confirm,
            "message": f"自动确认发货当前{'开启' if auto_confirm else '关闭'}"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/cookies/{cid}/remark")
def update_cookie_remark(cid: str, update_data: RemarkUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号备注"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 更新备注
        success = db_manager.update_cookie_remark(cid, update_data.remark)
        if success:
            log_with_user('info', f"更新账号备注: {cid} -> {update_data.remark}", current_user)
            return {
                "message": "备注更新成功",
                "remark": update_data.remark
            }
        else:
            raise HTTPException(status_code=500, detail="备注更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/cookies/{cid}/remark")
def get_cookie_remark(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号备注"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取Cookie详细信息（包含备注）
        cookie_details = db_manager.get_cookie_details(cid)
        if cookie_details:
            return {
                "remark": cookie_details.get('remark', ''),
                "message": "获取备注成功"
            }
        else:
            raise HTTPException(status_code=404, detail="账号不存在")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/cookies/{cid}/pause-duration")
def update_cookie_pause_duration(cid: str, update_data: PauseDurationUpdate, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新账号自动回复暂停时间"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 验证暂停时间范围（0-60分钟，0表示不暂停）
        if not (0 <= update_data.pause_duration <= 60):
            raise HTTPException(status_code=400, detail="暂停时间必须在0-60分钟之间（0表示不暂停）")

        # 更新暂停时间
        success = db_manager.update_cookie_pause_duration(cid, update_data.pause_duration)
        if success:
            log_with_user('info', f"更新账号自动回复暂停时间: {cid} -> {update_data.pause_duration}分钟", current_user)
            return {
                "message": "暂停时间更新成功",
                "pause_duration": update_data.pause_duration
            }
        else:
            raise HTTPException(status_code=500, detail="暂停时间更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/cookies/{cid}/pause-duration")
def get_cookie_pause_duration(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取账号自动回复暂停时间"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cid not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 获取暂停时间
        pause_duration = db_manager.get_cookie_pause_duration(cid)
        return {
            "pause_duration": pause_duration,
            "message": "获取暂停时间成功"
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))




class KeywordIn(BaseModel):
    keywords: Dict[str, str]  # key -> reply

class KeywordWithItemIdIn(BaseModel):
    keywords: List[Dict[str, Any]]  # [{"keyword": str, "reply": str, "item_id": str}]


@router.get("/keywords/{cid}")
def get_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from app.repositories.db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 直接从数据库获取所有关键词（避免重复计算）
    item_keywords = db_manager.get_keywords_with_item_id(cid)

    # 转换为统一格式
    all_keywords = []
    for keyword, reply, item_id in item_keywords:
        all_keywords.append({
            "keyword": keyword,
            "reply": reply,
            "item_id": item_id,
            "type": "item" if item_id else "normal"
        })

    return all_keywords


@router.get("/keywords-with-item-id/{cid}")
def get_keywords_with_item_id(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from app.repositories.db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 获取包含类型信息的关键词
    keywords = db_manager.get_keywords_with_type(cid)

    # 转换为前端需要的格式
    result = []
    for keyword_data in keywords:
        result.append({
            "keyword": keyword_data['keyword'],
            "reply": keyword_data['reply'],
            "item_id": keyword_data['item_id'] or "",
            "type": keyword_data['type'],
            "image_url": keyword_data['image_url']
        })

    return result


@router.post("/keywords/{cid}")
def update_keywords(cid: str, body: KeywordIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from app.repositories.db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    kw_list = [(k, v) for k, v in body.keywords.items()]
    log_with_user('info', f"更新Cookie关键字: {cid}, 数量: {len(kw_list)}", current_user)

    cookie_manager.manager.update_keywords(cid, kw_list)
    log_with_user('info', f"Cookie关键字更新成功: {cid}", current_user)
    return {"msg": "updated", "count": len(kw_list)}


@router.post("/keywords-with-item-id/{cid}")
def update_keywords_with_item_id(cid: str, body: KeywordWithItemIdIn, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新包含商品ID的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from app.repositories.db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        log_with_user('warning', f"尝试操作其他用户的Cookie关键字: {cid}", current_user)
        raise HTTPException(status_code=403, detail="无权限操作该Cookie")

    # 验证数据格式
    keywords_to_save = []
    keyword_set = set()  # 用于检查当前提交的关键词中是否有重复

    for kw_data in body.keywords:
        keyword = kw_data.get('keyword', '').strip()
        reply = kw_data.get('reply', '').strip()
        item_id = kw_data.get('item_id', '').strip() or None

        if not keyword:
            raise HTTPException(status_code=400, detail="关键词不能为空")

        # 检查当前提交的关键词中是否有重复
        keyword_key = f"{keyword}|{item_id or ''}"
        if keyword_key in keyword_set:
            item_id_text = f"（商品ID: {item_id}）" if item_id else "（通用关键词）"
            raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' {item_id_text} 在当前提交中重复")
        keyword_set.add(keyword_key)

        keywords_to_save.append((keyword, reply, item_id))

    # 保存关键词（只保存文本关键词，保留图片关键词）
    try:
        success = db_manager.save_text_keywords_only(cid, keywords_to_save)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词失败")
    except Exception as e:
        error_msg = str(e)

        # 检查是否是图片关键词冲突
        if "已存在（图片关键词）" in error_msg:
            # 直接使用数据库管理器提供的友好错误信息
            raise HTTPException(status_code=400, detail=error_msg)
        elif "UNIQUE constraint failed" in error_msg or "唯一约束冲突" in error_msg:
            # 尝试从错误信息中提取具体的冲突关键词
            conflict_keyword = None
            conflict_type = None

            # 检查是否是数据库管理器抛出的详细错误
            if "关键词唯一约束冲突" in error_msg:
                # 解析详细错误信息：关键词唯一约束冲突: Cookie=xxx, 关键词='xxx', 通用关键词/商品ID: xxx
                import re
                keyword_match = re.search(r"关键词='([^']+)'", error_msg)
                if keyword_match:
                    conflict_keyword = keyword_match.group(1)

                if "通用关键词" in error_msg:
                    conflict_type = "通用关键词"
                elif "商品ID:" in error_msg:
                    item_match = re.search(r"商品ID: ([^\s,]+)", error_msg)
                    if item_match:
                        conflict_type = f"商品关键词（商品ID: {item_match.group(1)}）"

            # 构造用户友好的错误信息
            if conflict_keyword and conflict_type:
                detail_msg = f'关键词 "{conflict_keyword}" （{conflict_type}） 已存在，请使用其他关键词或商品ID'
            elif "keywords.cookie_id, keywords.keyword" in error_msg:
                detail_msg = "关键词重复！该关键词已存在（可能是图片关键词或文本关键词），请使用其他关键词"
            else:
                detail_msg = "关键词重复！请使用不同的关键词或商品ID组合"

            raise HTTPException(status_code=400, detail=detail_msg)
        else:
            log_with_user('error', f"保存关键词时发生未知错误: {error_msg}", current_user)
            raise HTTPException(status_code=500, detail="保存关键词失败")

    log_with_user('info', f"更新Cookie关键字(含商品ID): {cid}, 数量: {len(keywords_to_save)}", current_user)
    return {"msg": "updated", "count": len(keywords_to_save)}


@router.get("/items/{cid}")
def get_items_list(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的商品列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from app.repositories.db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    try:
        # 获取该账号的所有商品
        with db_manager.lock:
            cursor = db_manager.conn.cursor()
            cursor.execute('''
            SELECT item_id, item_title, item_price, created_at
            FROM item_info
            WHERE cookie_id = ?
            ORDER BY created_at DESC
            ''', (cid,))

            items = []
            for row in cursor.fetchall():
                items.append({
                    'item_id': row[0],
                    'item_title': row[1] or '未知商品',
                    'item_price': row[2] or '价格未知',
                    'created_at': row[3]
                })

            return {"items": items, "count": len(items)}

    except Exception as e:
        logger.error(f"获取商品列表失败: {e}")
        raise HTTPException(status_code=500, detail="获取商品列表失败")


@router.get("/keywords-export/{cid}")
def export_keywords(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """导出指定账号的关键词为Excel文件"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from app.repositories.db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    try:
        # 获取关键词数据（包含类型信息）
        keywords = db_manager.get_keywords_with_type(cid)

        # 创建DataFrame，只导出文本类型的关键词
        data = []
        for keyword_data in keywords:
            # 只导出文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                data.append({
                    '关键词': keyword_data['keyword'],
                    '商品ID': keyword_data['item_id'] or '',
                    '关键词内容': keyword_data['reply']
                })

        # 如果没有数据，创建空的DataFrame但保留列名（作为模板）
        if not data:
            df = pd.DataFrame(columns=['关键词', '商品ID', '关键词内容'])
        else:
            df = pd.DataFrame(data)

        # 创建Excel文件
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='关键词数据', index=False)

            # 如果是空模板，添加一些示例说明
            if data == []:
                worksheet = writer.sheets['关键词数据']
                # 添加示例数据作为注释（从第2行开始）
                worksheet['A2'] = '你好'
                worksheet['B2'] = ''
                worksheet['C2'] = '您好！欢迎咨询，有什么可以帮助您的吗？'

                worksheet['A3'] = '价格'
                worksheet['B3'] = '123456'
                worksheet['C3'] = '这个商品的价格是99元，现在有优惠活动哦！'

                worksheet['A4'] = '发货'
                worksheet['B4'] = ''
                worksheet['C4'] = '我们会在24小时内发货，请耐心等待。'

                # 设置示例行的样式（浅灰色背景）
                from openpyxl.styles import PatternFill
                gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                for row in range(2, 5):
                    for col in range(1, 4):
                        worksheet.cell(row=row, column=col).fill = gray_fill

        output.seek(0)

        # 生成文件名（使用URL编码处理中文）
        from urllib.parse import quote
        if not data:
            filename = f"keywords_template_{cid}_{int(time.time())}.xlsx"
        else:
            filename = f"keywords_{cid}_{int(time.time())}.xlsx"
        encoded_filename = quote(filename.encode('utf-8'))

        # 返回文件
        return StreamingResponse(
            io.BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
            }
        )

    except Exception as e:
        logger.error(f"导出关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"导出关键词失败: {str(e)}")


@router.post("/keywords-import/{cid}")
async def import_keywords(cid: str, file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """导入Excel文件中的关键词到指定账号"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    user_id = current_user['user_id']
    from app.repositories.db_manager import db_manager
    user_cookies = db_manager.get_all_cookies(user_id)

    if cid not in user_cookies:
        raise HTTPException(status_code=403, detail="无权限访问该Cookie")

    # 检查文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传Excel文件(.xlsx或.xls)")

    try:
        # 读取Excel文件
        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))

        # 检查必要的列
        required_columns = ['关键词', '商品ID', '关键词内容']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise HTTPException(status_code=400, detail=f"Excel文件缺少必要的列: {', '.join(missing_columns)}")

        # 获取现有的文本类型关键词（用于比较更新/新增）
        existing_keywords = db_manager.get_keywords_with_type(cid)
        existing_dict = {}
        for keyword_data in existing_keywords:
            # 只考虑文本类型的关键词
            if keyword_data.get('type', 'text') == 'text':
                keyword = keyword_data['keyword']
                reply = keyword_data['reply']
                item_id = keyword_data['item_id']
                key = f"{keyword}|{item_id or ''}"
                existing_dict[key] = (keyword, reply, item_id)

        # 处理导入数据
        import_data = []
        update_count = 0
        add_count = 0

        for index, row in df.iterrows():
            keyword = str(row['关键词']).strip()
            item_id = str(row['商品ID']).strip() if pd.notna(row['商品ID']) and str(row['商品ID']).strip() else None
            reply = str(row['关键词内容']).strip()

            if not keyword:
                continue  # 跳过没有关键词的行

            # 检查是否重复
            key = f"{keyword}|{item_id or ''}"
            if key in existing_dict:
                # 更新现有关键词
                update_count += 1
            else:
                # 新增关键词
                add_count += 1

            import_data.append((keyword, reply, item_id))

        if not import_data:
            raise HTTPException(status_code=400, detail="Excel文件中没有有效的关键词数据")

        # 保存到数据库（只影响文本关键词，保留图片关键词）
        success = db_manager.save_text_keywords_only(cid, import_data)
        if not success:
            raise HTTPException(status_code=500, detail="保存关键词到数据库失败")

        log_with_user('info', f"导入关键词成功: {cid}, 新增: {add_count}, 更新: {update_count}", current_user)

        return {
            "msg": "导入成功",
            "total": len(import_data),
            "added": add_count,
            "updated": update_count
        }

    except pd.errors.EmptyDataError:
        raise HTTPException(status_code=400, detail="Excel文件为空")
    except pd.errors.ParserError:
        raise HTTPException(status_code=400, detail="Excel文件格式错误")
    except Exception as e:
        logger.error(f"导入关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"导入关键词失败: {str(e)}")


@router.post("/keywords/{cid}/image")
async def add_image_keyword(
    cid: str,
    keyword: str = Form(...),
    item_id: str = Form(default=""),
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """添加图片关键词"""
    logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}")

    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查参数
    if not keyword or not keyword.strip():
        raise HTTPException(status_code=400, detail="关键词不能为空")

    if not image or not image.filename:
        raise HTTPException(status_code=400, detail="请选择图片文件")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        logger.info(f"接收到图片关键词添加请求: cid={cid}, keyword={keyword}, item_id={item_id}, filename={image.filename}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片保存成功: {image_url}")

        # 先检查关键词是否已存在
        normalized_item_id = item_id if item_id and item_id.strip() else None
        if db_manager.check_keyword_duplicate(cid, keyword, normalized_item_id):
            # 删除已保存的图片
            image_manager.delete_image(image_url)
            if normalized_item_id:
                raise HTTPException(status_code=400, detail=f"关键词 '{keyword}' 在商品 '{normalized_item_id}' 中已存在")
            else:
                raise HTTPException(status_code=400, detail=f"通用关键词 '{keyword}' 已存在")

        # 保存图片关键词到数据库
        success = db_manager.save_image_keyword(cid, keyword, image_url, item_id or None)
        if not success:
            # 如果数据库保存失败，删除已保存的图片
            logger.error("数据库保存失败，删除已保存的图片")
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=400, detail="图片关键词保存失败，请稍后重试")

        log_with_user('info', f"添加图片关键词成功: {cid}, 关键词: {keyword}", current_user)

        return {
            "msg": "图片关键词添加成功",
            "keyword": keyword,
            "image_url": image_url,
            "item_id": item_id or None
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"添加图片关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"添加图片关键词失败: {str(e)}")


@router.post("/upload-image")
async def upload_image(
    image: UploadFile = File(...),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """上传图片（用于卡券等功能）"""
    try:
        logger.info(f"接收到图片上传请求: filename={image.filename}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片上传成功: {image_url}")

        return {
            "message": "图片上传成功",
            "image_url": image_url
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"图片上传失败: {e}")
        raise HTTPException(status_code=500, detail=f"图片上传失败: {str(e)}")


@router.get("/keywords-with-type/{cid}")
def get_keywords_with_type(cid: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取包含类型信息的关键词列表"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        keywords = db_manager.get_keywords_with_type(cid)
        return keywords
    except Exception as e:
        logger.error(f"获取关键词列表失败: {e}")
        raise HTTPException(status_code=500, detail=f"获取关键词列表失败: {str(e)}")


@router.delete("/keywords/{cid}/{index}")
def delete_keyword_by_index(cid: str, index: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """根据索引删除关键词"""
    if cookie_manager.manager is None:
        raise HTTPException(status_code=500, detail="CookieManager 未就绪")

    # 检查cookie是否属于当前用户
    cookie_details = db_manager.get_cookie_details(cid)
    if not cookie_details or cookie_details['user_id'] != current_user['user_id']:
        raise HTTPException(status_code=404, detail="账号不存在或无权限")

    try:
        # 先获取要删除的关键词信息（用于删除图片文件）
        keywords = db_manager.get_keywords_with_type(cid)
        if 0 <= index < len(keywords):
            keyword_data = keywords[index]

            # 删除关键词
            success = db_manager.delete_keyword_by_index(cid, index)
            if not success:
                raise HTTPException(status_code=400, detail="删除关键词失败")

            # 如果是图片关键词，删除对应的图片文件
            if keyword_data.get('type') == 'image' and keyword_data.get('image_url'):
                image_manager.delete_image(keyword_data['image_url'])

            log_with_user('info', f"删除关键词成功: {cid}, 索引: {index}, 关键词: {keyword_data.get('keyword')}", current_user)

            return {"msg": "删除成功"}
        else:
            raise HTTPException(status_code=400, detail="关键词索引无效")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除关键词失败: {e}")
        raise HTTPException(status_code=500, detail=f"删除关键词失败: {str(e)}")


@router.get("/debug/keywords-table-info")
def debug_keywords_table_info(current_user: Dict[str, Any] = Depends(get_current_user)):
    """调试：检查keywords表结构"""
    try:
        conn = sqlite3.connect(db_manager.db_path)
        cursor = conn.cursor()

        # 获取表结构信息
        cursor.execute("PRAGMA table_info(keywords)")
        columns = cursor.fetchall()

        # 获取数据库版本
        cursor.execute("SELECT value FROM system_settings WHERE key = 'db_version'")
        version_result = cursor.fetchone()
        db_version = version_result[0] if version_result else "未知"

        conn.close()

        return {
            "db_version": db_version,
            "table_columns": [{"name": col[1], "type": col[2], "default": col[4]} for col in columns]
        }
    except Exception as e:
        logger.error(f"检查表结构失败: {e}")
        raise HTTPException(status_code=500, detail=f"检查表结构失败: {str(e)}")


# 卡券管理API
@router.get("/cards")
def get_cards(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的卡券列表"""
    try:
        from app.repositories.db_manager import db_manager
        user_id = current_user['user_id']
        cards = db_manager.get_all_cards(user_id)
        return cards
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/cards")
def create_card(card_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建新卡券"""
    try:
        from app.repositories.db_manager import db_manager
        user_id = current_user['user_id']
        card_name = card_data.get('name', '未命名卡券')

        log_with_user('info', f"创建卡券: {card_name}", current_user)

        # 验证多规格字段
        is_multi_spec = card_data.get('is_multi_spec', False)
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        card_id = db_manager.create_card(
            name=card_data.get('name'),
            card_type=card_data.get('type'),
            api_config=card_data.get('api_config'),
            text_content=card_data.get('text_content'),
            data_content=card_data.get('data_content'),
            image_url=card_data.get('image_url'),
            description=card_data.get('description'),
            enabled=card_data.get('enabled', True),
            delay_seconds=card_data.get('delay_seconds', 0),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name') if is_multi_spec else None,
            spec_value=card_data.get('spec_value') if is_multi_spec else None,
            user_id=user_id
        )

        log_with_user('info', f"卡券创建成功: {card_name} (ID: {card_id})", current_user)
        return {"id": card_id, "message": "卡券创建成功"}
    except Exception as e:
        log_with_user('error', f"创建卡券失败: {card_data.get('name', '未知')} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/cards/{card_id}")
def get_card(card_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取单个卡券详情"""
    try:
        from app.repositories.db_manager import db_manager
        user_id = current_user['user_id']
        card = db_manager.get_card_by_id(card_id, user_id)
        if card:
            return card
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/cards/{card_id}")
def update_card(card_id: int, card_data: dict, _: None = Depends(require_auth)):
    """更新卡券"""
    try:
        from app.repositories.db_manager import db_manager
        # 验证多规格字段
        is_multi_spec = card_data.get('is_multi_spec')
        if is_multi_spec:
            if not card_data.get('spec_name') or not card_data.get('spec_value'):
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        success = db_manager.update_card(
            card_id=card_id,
            name=card_data.get('name'),
            card_type=card_data.get('type'),
            api_config=card_data.get('api_config'),
            text_content=card_data.get('text_content'),
            data_content=card_data.get('data_content'),
            image_url=card_data.get('image_url'),
            description=card_data.get('description'),
            enabled=card_data.get('enabled', True),
            delay_seconds=card_data.get('delay_seconds'),
            is_multi_spec=is_multi_spec,
            spec_name=card_data.get('spec_name'),
            spec_value=card_data.get('spec_value')
        )
        if success:
            return {"message": "卡券更新成功"}
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/cards/{card_id}/image")
async def update_card_with_image(
    card_id: int,
    image: UploadFile = File(...),
    name: str = Form(...),
    type: str = Form(...),
    description: str = Form(default=""),
    delay_seconds: int = Form(default=0),
    enabled: bool = Form(default=True),
    is_multi_spec: bool = Form(default=False),
    spec_name: str = Form(default=""),
    spec_value: str = Form(default=""),
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新带图片的卡券"""
    try:
        logger.info(f"接收到带图片的卡券更新请求: card_id={card_id}, name={name}, type={type}")

        # 验证图片文件
        if not image.content_type or not image.content_type.startswith('image/'):
            logger.warning(f"无效的图片文件类型: {image.content_type}")
            raise HTTPException(status_code=400, detail="请上传图片文件")

        # 验证多规格字段
        if is_multi_spec:
            if not spec_name or not spec_value:
                raise HTTPException(status_code=400, detail="多规格卡券必须提供规格名称和规格值")

        # 读取图片数据
        image_data = await image.read()
        logger.info(f"读取图片数据成功，大小: {len(image_data)} bytes")

        # 保存图片
        image_url = image_manager.save_image(image_data, image.filename)
        if not image_url:
            logger.error("图片保存失败")
            raise HTTPException(status_code=400, detail="图片保存失败")

        logger.info(f"图片保存成功: {image_url}")

        # 更新卡券
        from app.repositories.db_manager import db_manager
        success = db_manager.update_card(
            card_id=card_id,
            name=name,
            card_type=type,
            image_url=image_url,
            description=description,
            enabled=enabled,
            delay_seconds=delay_seconds,
            is_multi_spec=is_multi_spec,
            spec_name=spec_name if is_multi_spec else None,
            spec_value=spec_value if is_multi_spec else None
        )

        if success:
            logger.info(f"卡券更新成功: {name} (ID: {card_id})")
            return {"message": "卡券更新成功", "image_url": image_url}
        else:
            # 如果数据库更新失败，删除已保存的图片
            image_manager.delete_image(image_url)
            raise HTTPException(status_code=404, detail="卡券不存在")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新带图片的卡券失败: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# 自动发货规则API
@router.get("/delivery-rules")
def get_delivery_rules(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取发货规则列表"""
    try:
        from app.repositories.db_manager import db_manager
        user_id = current_user['user_id']
        rules = db_manager.get_all_delivery_rules(user_id)
        return rules
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/delivery-rules")
def create_delivery_rule(rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """创建新发货规则"""
    try:
        from app.repositories.db_manager import db_manager
        user_id = current_user['user_id']
        rule_id = db_manager.create_delivery_rule(
            keyword=rule_data.get('keyword'),
            card_id=rule_data.get('card_id'),
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        return {"id": rule_id, "message": "发货规则创建成功"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/delivery-rules/{rule_id}")
def get_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取单个发货规则详情"""
    try:
        from app.repositories.db_manager import db_manager
        user_id = current_user['user_id']
        rule = db_manager.get_delivery_rule_by_id(rule_id, user_id)
        if rule:
            return rule
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/delivery-rules/{rule_id}")
def update_delivery_rule(rule_id: int, rule_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新发货规则"""
    try:
        from app.repositories.db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.update_delivery_rule(
            rule_id=rule_id,
            keyword=rule_data.get('keyword'),
            card_id=rule_data.get('card_id'),
            delivery_count=rule_data.get('delivery_count', 1),
            enabled=rule_data.get('enabled', True),
            description=rule_data.get('description'),
            user_id=user_id
        )
        if success:
            return {"message": "发货规则更新成功"}
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/cards/{card_id}")
def delete_card(card_id: int, _: None = Depends(require_auth)):
    """删除卡券"""
    try:
        from app.repositories.db_manager import db_manager
        success = db_manager.delete_card(card_id)
        if success:
            return {"message": "卡券删除成功"}
        else:
            raise HTTPException(status_code=404, detail="卡券不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/delivery-rules/{rule_id}")
def delete_delivery_rule(rule_id: int, current_user: Dict[str, Any] = Depends(get_current_user)):
    """删除发货规则"""
    try:
        from app.repositories.db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.delete_delivery_rule(rule_id, user_id)
        if success:
            return {"message": "发货规则删除成功"}
        else:
            raise HTTPException(status_code=404, detail="发货规则不存在")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ==================== 备份和恢复 API ====================

@router.get("/backup/export")
def export_backup(current_user: Dict[str, Any] = Depends(get_current_user)):
    """导出用户备份"""
    try:
        from app.repositories.db_manager import db_manager
        user_id = current_user['user_id']
        username = current_user['username']

        # 导出当前用户的数据
        backup_data = db_manager.export_backup(user_id)

        # 生成文件名
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"xianyu_backup_{username}_{timestamp}.json"

        # 返回JSON响应，设置下载头
        response = JSONResponse(content=backup_data)
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/json"

        return response
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出备份失败: {str(e)}")


@router.post("/backup/import")
def import_backup(file: UploadFile = File(...), current_user: Dict[str, Any] = Depends(get_current_user)):
    """导入用户备份"""
    try:
        # 验证文件类型
        if not file.filename.endswith('.json'):
            raise HTTPException(status_code=400, detail="只支持JSON格式的备份文件")

        # 读取文件内容
        content = file.file.read()
        backup_data = json.loads(content.decode('utf-8'))

        # 导入备份到当前用户
        from app.repositories.db_manager import db_manager
        user_id = current_user['user_id']
        success = db_manager.import_backup(backup_data, user_id)

        if success:
            # 备份导入成功后，刷新 CookieManager 的内存缓存
            from app.services import cookie_manager
            if cookie_manager.manager:
                try:
                    cookie_manager.manager.reload_from_db()
                    logger.info("备份导入后已刷新 CookieManager 缓存")
                except Exception as e:
                    logger.error(f"刷新 CookieManager 缓存失败: {e}")

            return {"message": "备份导入成功"}
        else:
            raise HTTPException(status_code=400, detail="备份导入失败")

    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="备份文件格式无效")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导入备份失败: {str(e)}")


@router.post("/system/reload-cache")
def reload_cache(_: None = Depends(require_auth)):
    """重新加载系统缓存（用于手动刷新数据）"""
    try:
        from app.services import cookie_manager
        if cookie_manager.manager:
            success = cookie_manager.manager.reload_from_db()
            if success:
                return {"message": "系统缓存已刷新", "success": True}
            else:
                raise HTTPException(status_code=500, detail="缓存刷新失败")
        else:
            raise HTTPException(status_code=500, detail="CookieManager 未初始化")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"刷新缓存失败: {str(e)}")


# ==================== 商品管理 API ====================

@router.get("/items")
def get_all_items(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的所有商品信息"""
    try:
        # 只返回当前用户的商品信息
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_items_by_cookie(cookie_id)
            all_items.extend(items)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")


# ==================== 商品搜索 API ====================

class ItemSearchRequest(BaseModel):
    keyword: str
    page: int = 1
    page_size: int = 20

class ItemSearchMultipleRequest(BaseModel):
    keyword: str
    total_pages: int = 1

@router.post("/items/search")
async def search_items(
    search_request: ItemSearchRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """搜索闲鱼商品"""
    user_info = f"【{current_user.get('username', 'unknown')}#{current_user.get('user_id', 'unknown')}】" if current_user else "【未登录】"

    try:
        logger.info(f"{user_info} 开始单页搜索: 关键词='{search_request.keyword}', 页码={search_request.page}, 每页={search_request.page_size}")

        from app.utils.item_search import search_xianyu_items

        # 执行搜索
        result = await search_xianyu_items(
            keyword=search_request.keyword,
            page=search_request.page,
            page_size=search_request.page_size
        )

        # 检查是否有错误
        has_error = result.get("error")
        items_count = len(result.get("items", []))

        logger.info(f"{user_info} 单页搜索完成: 获取到 {items_count} 条数据" +
                   (f", 错误: {has_error}" if has_error else ""))

        response_data = {
            "success": True,
            "data": result.get("items", []),
            "total": result.get("total", 0),
            "page": search_request.page,
            "page_size": search_request.page_size,
            "keyword": search_request.keyword,
            "is_real_data": result.get("is_real_data", False),
            "source": result.get("source", "unknown")
        }

        # 如果有错误信息，也包含在响应中
        if has_error:
            response_data["error"] = has_error

        return response_data

    except Exception as e:
        error_msg = str(e)
        logger.error(f"{user_info} 商品搜索失败: {error_msg}")
        raise HTTPException(status_code=500, detail=f"商品搜索失败: {error_msg}")


@router.get("/cookies/check")
async def check_valid_cookies(
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """检查是否有有效的cookies账户（必须是启用状态）"""
    try:
        if cookie_manager.manager is None:
            return {
                "success": True,
                "hasValidCookies": False,
                "validCount": 0,
                "enabledCount": 0,
                "totalCount": 0
            }

        from app.repositories.db_manager import db_manager

        # 获取所有cookies
        all_cookies = db_manager.get_all_cookies()

        # 检查启用状态和有效性
        valid_cookies = []
        enabled_cookies = []

        for cookie_id, cookie_value in all_cookies.items():
            # 检查是否启用
            is_enabled = cookie_manager.manager.get_cookie_status(cookie_id)
            if is_enabled:
                enabled_cookies.append(cookie_id)
                # 检查是否有效（长度大于50）
                if len(cookie_value) > 50:
                    valid_cookies.append(cookie_id)

        return {
            "success": True,
            "hasValidCookies": len(valid_cookies) > 0,
            "validCount": len(valid_cookies),
            "enabledCount": len(enabled_cookies),
            "totalCount": len(all_cookies)
        }

    except Exception as e:
        logger.error(f"检查cookies失败: {str(e)}")
        return {
            "success": False,
            "hasValidCookies": False,
            "error": str(e)
        }

@router.post("/items/search_multiple")
async def search_multiple_pages(
    search_request: ItemSearchMultipleRequest,
    current_user: Optional[Dict[str, Any]] = Depends(get_current_user_optional)
):
    """搜索多页闲鱼商品"""
    user_info = f"【{current_user.get('username', 'unknown')}#{current_user.get('user_id', 'unknown')}】" if current_user else "【未登录】"

    try:
        logger.info(f"{user_info} 开始多页搜索: 关键词='{search_request.keyword}', 页数={search_request.total_pages}")

        from app.utils.item_search import search_multiple_pages_xianyu

        # 执行多页搜索
        result = await search_multiple_pages_xianyu(
            keyword=search_request.keyword,
            total_pages=search_request.total_pages
        )

        # 检查是否有错误
        has_error = result.get("error")
        items_count = len(result.get("items", []))

        logger.info(f"{user_info} 多页搜索完成: 获取到 {items_count} 条数据" +
                   (f", 错误: {has_error}" if has_error else ""))

        response_data = {
            "success": True,
            "data": result.get("items", []),
            "total": result.get("total", 0),
            "total_pages": search_request.total_pages,
            "keyword": search_request.keyword,
            "is_real_data": result.get("is_real_data", False),
            "is_fallback": result.get("is_fallback", False),
            "source": result.get("source", "unknown")
        }

        # 如果有错误信息，也包含在响应中
        if has_error:
            response_data["error"] = has_error

        return response_data

    except Exception as e:
        error_msg = str(e)
        logger.error(f"{user_info} 多页商品搜索失败: {error_msg}")
        raise HTTPException(status_code=500, detail=f"多页商品搜索失败: {error_msg}")



@router.get("/items/cookie/{cookie_id}")
def get_items_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定Cookie的商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        items = db_manager.get_items_by_cookie(cookie_id)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")


@router.get("/items/{cookie_id}/{item_id}")
def get_item_detail(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取商品详情"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        item = db_manager.get_item_info(cookie_id, item_id)
        if not item:
            raise HTTPException(status_code=404, detail="商品不存在")
        return {"item": item}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品详情失败: {str(e)}")


class ItemDetailUpdate(BaseModel):
    item_detail: str


@router.put("/items/{cookie_id}/{item_id}")
def update_item_detail(
    cookie_id: str,
    item_id: str,
    update_data: ItemDetailUpdate,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """更新商品详情"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.update_item_detail(cookie_id, item_id, update_data.item_detail)
        if success:
            return {"message": "商品详情更新成功"}
        else:
            raise HTTPException(status_code=400, detail="更新失败")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新商品详情失败: {str(e)}")


@router.delete("/items/{cookie_id}/{item_id}")
def delete_item_info(
    cookie_id: str,
    item_id: str,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """删除商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        success = db_manager.delete_item_info(cookie_id, item_id)
        if success:
            return {"message": "商品信息删除成功"}
        else:
            raise HTTPException(status_code=404, detail="商品信息不存在")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除商品信息异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


class BatchDeleteRequest(BaseModel):
    items: List[dict]  # [{"cookie_id": "xxx", "item_id": "yyy"}, ...]


class AIReplySettings(BaseModel):
    ai_enabled: bool
    model_name: str = "qwen-plus"
    api_key: str = ""
    base_url: str = "https://dashscope.aliyuncs.com/compatible-mode/v1"
    max_discount_percent: int = 10
    max_discount_amount: int = 100
    max_bargain_rounds: int = 3
    custom_prompts: str = ""


@router.delete("/items/batch")
def batch_delete_items(
    request: BatchDeleteRequest,
    _: None = Depends(require_auth)
):
    """批量删除商品信息"""
    try:
        if not request.items:
            raise HTTPException(status_code=400, detail="删除列表不能为空")

        success_count = db_manager.batch_delete_item_info(request.items)
        total_count = len(request.items)

        return {
            "message": "批量删除完成",
            "success_count": success_count,
            "total_count": total_count,
            "failed_count": total_count - success_count
        }
    except Exception as e:
        logger.error(f"批量删除商品信息异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


# ==================== AI回复管理API ====================

@router.get("/ai-reply-settings/{cookie_id}")
def get_ai_reply_settings(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定账号的AI回复设置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        settings = db_manager.get_ai_reply_settings(cookie_id)
        return settings
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@router.put("/ai-reply-settings/{cookie_id}")
def update_ai_reply_settings(cookie_id: str, settings: AIReplySettings, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新指定账号的AI回复设置"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限操作该Cookie")

        # 检查账号是否存在
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager 未就绪')

        # 保存设置
        settings_dict = settings.dict()
        success = db_manager.save_ai_reply_settings(cookie_id, settings_dict)

        if success:
            # 清理客户端缓存，强制重新创建
            ai_reply_engine.clear_client_cache(cookie_id)

            # 如果启用了AI回复，记录日志
            if settings.ai_enabled:
                logger.info(f"账号 {cookie_id} 启用AI回复")
            else:
                logger.info(f"账号 {cookie_id} 禁用AI回复")

            return {"message": "AI回复设置更新成功"}
        else:
            raise HTTPException(status_code=400, detail="更新失败")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"更新AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@router.get("/ai-reply-settings")
def get_all_ai_reply_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户所有账号的AI回复设置"""
    try:
        # 只返回当前用户的AI回复设置
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_settings = db_manager.get_all_ai_reply_settings()
        # 过滤只属于当前用户的设置
        user_settings = {cid: settings for cid, settings in all_settings.items() if cid in user_cookies}
        return user_settings
    except Exception as e:
        logger.error(f"获取所有AI回复设置异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


@router.post("/ai-reply-test/{cookie_id}")
def test_ai_reply(cookie_id: str, test_data: dict, _: None = Depends(require_auth)):
    """测试AI回复功能"""
    try:
        # 检查账号是否存在
        if cookie_manager.manager is None:
            raise HTTPException(status_code=500, detail='CookieManager 未就绪')

        if cookie_id not in cookie_manager.manager.cookies:
            raise HTTPException(status_code=404, detail='账号不存在')

        # 检查是否启用AI回复
        if not ai_reply_engine.is_ai_enabled(cookie_id):
            raise HTTPException(status_code=400, detail='该账号未启用AI回复')

        # 构造测试数据
        test_message = test_data.get('message', '你好')
        test_item_info = {
            'title': test_data.get('item_title', '测试商品'),
            'price': test_data.get('item_price', 100),
            'desc': test_data.get('item_desc', '这是一个测试商品')
        }

        # 生成测试回复
        reply = ai_reply_engine.generate_reply(
            message=test_message,
            item_info=test_item_info,
            chat_id=f"test_{int(time.time())}",
            cookie_id=cookie_id,
            user_id="test_user",
            item_id="test_item"
        )

        if reply:
            return {"message": "测试成功", "reply": reply}
        else:
            raise HTTPException(status_code=400, detail="AI回复生成失败")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"测试AI回复异常: {e}")
        raise HTTPException(status_code=500, detail=f"服务器错误: {str(e)}")


# ==================== 日志管理API ====================

@router.get("/logs")
async def get_logs(lines: int = 200, level: str = None, source: str = None, _: None = Depends(require_auth)):
    """获取实时系统日志"""
    try:
        # 获取文件日志收集器
        collector = get_file_log_collector()

        # 获取日志
        logs = collector.get_logs(lines=lines, level_filter=level, source_filter=source)

        return {"success": True, "logs": logs}

    except Exception as e:
        return {"success": False, "message": f"获取日志失败: {str(e)}", "logs": []}


@router.get("/risk-control-logs")
async def get_risk_control_logs(
    cookie_id: str = None,
    limit: int = 100,
    offset: int = 0,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """获取风控日志（管理员专用）"""
    try:
        log_with_user('info', f"查询风控日志: cookie_id={cookie_id}, limit={limit}, offset={offset}", admin_user)

        # 获取风控日志
        logs = db_manager.get_risk_control_logs(cookie_id=cookie_id, limit=limit, offset=offset)
        total_count = db_manager.get_risk_control_logs_count(cookie_id=cookie_id)

        log_with_user('info', f"风控日志查询成功，共 {len(logs)} 条记录，总计 {total_count} 条", admin_user)

        return {
            "success": True,
            "data": logs,
            "total": total_count,
            "limit": limit,
            "offset": offset
        }

    except Exception as e:
        log_with_user('error', f"获取风控日志失败: {str(e)}", admin_user)
        return {
            "success": False,
            "message": f"获取风控日志失败: {str(e)}",
            "data": [],
            "total": 0
        }


@router.delete("/risk-control-logs/{log_id}")
async def delete_risk_control_log(
    log_id: int,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """删除风控日志记录（管理员专用）"""
    try:
        log_with_user('info', f"删除风控日志记录: {log_id}", admin_user)

        success = db_manager.delete_risk_control_log(log_id)

        if success:
            log_with_user('info', f"风控日志删除成功: {log_id}", admin_user)
            return {"success": True, "message": "删除成功"}
        else:
            log_with_user('warning', f"风控日志删除失败: {log_id}", admin_user)
            return {"success": False, "message": "删除失败，记录可能不存在"}

    except Exception as e:
        log_with_user('error', f"删除风控日志失败: {log_id} - {str(e)}", admin_user)
        return {"success": False, "message": f"删除失败: {str(e)}"}


@router.get("/logs/stats")
async def get_log_stats(_: None = Depends(require_auth)):
    """获取日志统计信息"""
    try:
        collector = get_file_log_collector()
        stats = collector.get_stats()

        return {"success": True, "stats": stats}

    except Exception as e:
        return {"success": False, "message": f"获取日志统计失败: {str(e)}", "stats": {}}


@router.post("/logs/clear")
async def clear_logs(_: None = Depends(require_auth)):
    """清空日志"""
    try:
        collector = get_file_log_collector()
        collector.clear_logs()

        return {"success": True, "message": "日志已清空"}

    except Exception as e:
        return {"success": False, "message": f"清空日志失败: {str(e)}"}


# ==================== 商品管理API ====================

@router.post("/items/get-all-from-account")
async def get_all_items_from_account(request: dict, _: None = Depends(require_auth)):
    """从指定账号获取所有商品信息"""
    try:
        cookie_id = request.get('cookie_id')
        if not cookie_id:
            return {"success": False, "message": "缺少cookie_id参数"}

        # 获取指定账号的cookie信息
        cookie_info = db_manager.get_cookie_by_id(cookie_id)
        if not cookie_info:
            return {"success": False, "message": "未找到指定的账号信息"}

        cookies_str = cookie_info.get('cookies_str', '')
        if not cookies_str:
            return {"success": False, "message": "账号cookie信息为空"}

        # 创建XianyuLive实例，传入正确的cookie_id
        from app.services.xianyu_async import XianyuLive
        xianyu_instance = XianyuLive(cookies_str, cookie_id)

        # 调用获取所有商品信息的方法（自动分页）
        logger.info(f"开始获取账号 {cookie_id} 的所有商品信息")
        result = await xianyu_instance.get_all_items()

        # 关闭session
        await xianyu_instance.close_session()

        if result.get('error'):
            logger.error(f"获取商品信息失败: {result['error']}")
            return {"success": False, "message": result['error']}
        else:
            total_count = result.get('total_count', 0)
            total_pages = result.get('total_pages', 1)
            logger.info(f"成功获取账号 {cookie_id} 的 {total_count} 个商品（共{total_pages}页）")
            return {
                "success": True,
                "message": f"成功获取 {total_count} 个商品（共{total_pages}页），详细信息已打印到控制台",
                "total_count": total_count,
                "total_pages": total_pages
            }

    except Exception as e:
        logger.error(f"获取账号商品信息异常: {str(e)}")
        return {"success": False, "message": f"获取商品信息异常: {str(e)}"}


@router.post("/items/get-by-page")
async def get_items_by_page(request: dict, _: None = Depends(require_auth)):
    """从指定账号按页获取商品信息"""
    try:
        # 验证参数
        cookie_id = request.get('cookie_id')
        page_number = request.get('page_number', 1)
        page_size = request.get('page_size', 20)

        if not cookie_id:
            return {"success": False, "message": "缺少cookie_id参数"}

        # 验证分页参数
        try:
            page_number = int(page_number)
            page_size = int(page_size)
        except (ValueError, TypeError):
            return {"success": False, "message": "页码和每页数量必须是数字"}

        if page_number < 1:
            return {"success": False, "message": "页码必须大于0"}

        if page_size < 1 or page_size > 100:
            return {"success": False, "message": "每页数量必须在1-100之间"}

        # 获取账号信息
        account = db_manager.get_cookie_by_id(cookie_id)
        if not account:
            return {"success": False, "message": "账号不存在"}

        cookies_str = account['cookies_str']
        if not cookies_str:
            return {"success": False, "message": "账号cookies为空"}

        # 创建XianyuLive实例，传入正确的cookie_id
        from app.services.xianyu_async import XianyuLive
        xianyu_instance = XianyuLive(cookies_str, cookie_id)

        # 调用获取指定页商品信息的方法
        logger.info(f"开始获取账号 {cookie_id} 第{page_number}页商品信息（每页{page_size}条）")
        result = await xianyu_instance.get_item_list_info(page_number, page_size)

        # 关闭session
        await xianyu_instance.close_session()

        if result.get('error'):
            logger.error(f"获取商品信息失败: {result['error']}")
            return {"success": False, "message": result['error']}
        else:
            current_count = result.get('current_count', 0)
            logger.info(f"成功获取账号 {cookie_id} 第{page_number}页 {current_count} 个商品")
            return {
                "success": True,
                "message": f"成功获取第{page_number}页 {current_count} 个商品，详细信息已打印到控制台",
                "page_number": page_number,
                "page_size": page_size,
                "current_count": current_count
            }

    except Exception as e:
        logger.error(f"获取账号商品信息异常: {str(e)}")
        return {"success": False, "message": f"获取商品信息异常: {str(e)}"}


# ------------------------- 用户设置接口 -------------------------

@router.get('/user-settings')
def get_user_settings(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的设置"""
    from app.repositories.db_manager import db_manager
    try:
        user_id = current_user['user_id']
        settings = db_manager.get_user_settings(user_id)
        return settings
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put('/user-settings/{key}')
def update_user_setting(key: str, setting_data: dict, current_user: Dict[str, Any] = Depends(get_current_user)):
    """更新用户设置"""
    from app.repositories.db_manager import db_manager
    try:
        user_id = current_user['user_id']
        value = setting_data.get('value')
        description = setting_data.get('description', '')

        log_with_user('info', f"更新用户设置: {key} = {value}", current_user)

        success = db_manager.set_user_setting(user_id, key, value, description)
        if success:
            log_with_user('info', f"用户设置更新成功: {key}", current_user)
            return {'msg': 'setting updated', 'key': key, 'value': value}
        else:
            log_with_user('error', f"用户设置更新失败: {key}", current_user)
            raise HTTPException(status_code=400, detail='更新失败')
    except Exception as e:
        log_with_user('error', f"更新用户设置异常: {key} - {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=str(e))

@router.get('/user-settings/{key}')
def get_user_setting(key: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取用户特定设置"""
    from app.repositories.db_manager import db_manager
    try:
        user_id = current_user['user_id']
        setting = db_manager.get_user_setting(user_id, key)
        if setting:
            return setting
        else:
            raise HTTPException(status_code=404, detail='设置不存在')
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 管理员专用接口 -------------------------

@router.get('/admin/users')
def get_all_users(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取所有用户信息（管理员专用）"""
    from app.repositories.db_manager import db_manager
    try:
        log_with_user('info', "查询所有用户信息", admin_user)
        users = db_manager.get_all_users()

        # 为每个用户添加统计信息
        for user in users:
            user_id = user['id']
            # 统计用户的Cookie数量
            user_cookies = db_manager.get_all_cookies(user_id)
            user['cookie_count'] = len(user_cookies)

            # 统计用户的卡券数量
            user_cards = db_manager.get_all_cards(user_id)
            user['card_count'] = len(user_cards) if user_cards else 0

            # 隐藏密码字段
            if 'password_hash' in user:
                del user['password_hash']

        log_with_user('info', f"返回用户信息，共 {len(users)} 个用户", admin_user)
        return {"users": users}
    except Exception as e:
        log_with_user('error', f"获取用户信息失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@router.delete('/admin/users/{user_id}')
def delete_user(user_id: int, admin_user: Dict[str, Any] = Depends(require_admin)):
    """删除用户（管理员专用）"""
    from app.repositories.db_manager import db_manager
    try:
        # 不能删除管理员自己
        if user_id == admin_user['user_id']:
            log_with_user('warning', "尝试删除管理员自己", admin_user)
            raise HTTPException(status_code=400, detail="不能删除管理员自己")

        # 获取要删除的用户信息
        user_to_delete = db_manager.get_user_by_id(user_id)
        if not user_to_delete:
            raise HTTPException(status_code=404, detail="用户不存在")

        log_with_user('info', f"准备删除用户: {user_to_delete['username']} (ID: {user_id})", admin_user)

        # 删除用户及其相关数据
        success = db_manager.delete_user_and_data(user_id)

        if success:
            log_with_user('info', f"用户删除成功: {user_to_delete['username']} (ID: {user_id})", admin_user)
            return {"message": f"用户 {user_to_delete['username']} 删除成功"}
        else:
            log_with_user('error', f"用户删除失败: {user_to_delete['username']} (ID: {user_id})", admin_user)
            raise HTTPException(status_code=400, detail="删除失败")
    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除用户异常: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@router.get('/admin/risk-control-logs')
async def get_admin_risk_control_logs(
    cookie_id: str = None,
    limit: int = 100,
    offset: int = 0,
    admin_user: Dict[str, Any] = Depends(require_admin)
):
    """获取风控日志（管理员专用）"""
    try:
        log_with_user('info', f"查询风控日志: cookie_id={cookie_id}, limit={limit}, offset={offset}", admin_user)

        # 获取风控日志
        logs = db_manager.get_risk_control_logs(cookie_id=cookie_id, limit=limit, offset=offset)
        total_count = db_manager.get_risk_control_logs_count(cookie_id=cookie_id)

        log_with_user('info', f"风控日志查询成功，共 {len(logs)} 条记录，总计 {total_count} 条", admin_user)

        return {
            "success": True,
            "data": logs,
            "total": total_count,
            "limit": limit,
            "offset": offset
        }

    except Exception as e:
        log_with_user('error', f"查询风控日志失败: {str(e)}", admin_user)
        return {"success": False, "message": f"查询失败: {str(e)}", "data": [], "total": 0}


@router.get('/admin/cookies')
def get_admin_cookies(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取所有Cookie信息（管理员专用）"""
    try:
        log_with_user('info', "查询所有Cookie信息", admin_user)

        if cookie_manager.manager is None:
            return {
                "success": True,
                "cookies": [],
                "message": "CookieManager 未就绪"
            }

        # 获取所有用户的cookies
        from app.repositories.db_manager import db_manager
        all_users = db_manager.get_all_users()
        all_cookies = []

        for user in all_users:
            user_id = user['id']
            user_cookies = db_manager.get_all_cookies(user_id)
            for cookie_id, cookie_value in user_cookies.items():
                # 获取cookie详细信息
                cookie_details = db_manager.get_cookie_details(cookie_id)
                cookie_info = {
                    'cookie_id': cookie_id,
                    'user_id': user_id,
                    'username': user['username'],
                    'nickname': cookie_details.get('remark', '') if cookie_details else '',
                    'enabled': cookie_manager.manager.get_cookie_status(cookie_id)
                }
                all_cookies.append(cookie_info)

        log_with_user('info', f"获取到 {len(all_cookies)} 个Cookie", admin_user)
        return {
            "success": True,
            "cookies": all_cookies,
            "total": len(all_cookies)
        }

    except Exception as e:
        log_with_user('error', f"获取Cookie信息失败: {str(e)}", admin_user)
        return {
            "success": False,
            "cookies": [],
            "message": f"获取失败: {str(e)}"
        }


@router.get('/admin/logs')
def get_system_logs(admin_user: Dict[str, Any] = Depends(require_admin),
                   lines: int = 100,
                   level: str = None):
    """获取系统日志（管理员专用）"""
    try:
        log_with_user('info', f"查询系统日志，行数: {lines}, 级别: {level}", admin_user)

        # 查找日志文件
        log_files = glob.glob("logs/xianyu_*.log")
        logger.info(f"找到日志文件: {log_files}")

        if not log_files:
            logger.warning("未找到日志文件")
            return {"logs": [], "message": "未找到日志文件", "success": False}

        # 获取最新的日志文件
        latest_log_file = max(log_files, key=os.path.getctime)
        logger.info(f"使用最新日志文件: {latest_log_file}")

        logs = []
        try:
            with open(latest_log_file, 'r', encoding='utf-8') as f:
                all_lines = f.readlines()
                logger.info(f"读取到 {len(all_lines)} 行日志")

                # 如果指定了日志级别，进行过滤
                if level:
                    filtered_lines = [line for line in all_lines if f"| {level.upper()} |" in line]
                    logger.info(f"按级别 {level} 过滤后剩余 {len(filtered_lines)} 行")
                else:
                    filtered_lines = all_lines

                # 获取最后N行
                recent_lines = filtered_lines[-lines:] if len(filtered_lines) > lines else filtered_lines
                logger.info(f"取最后 {len(recent_lines)} 行日志")

                for line in recent_lines:
                    logs.append(line.strip())

        except Exception as e:
            logger.error(f"读取日志文件失败: {str(e)}")
            log_with_user('error', f"读取日志文件失败: {str(e)}", admin_user)
            return {"logs": [], "message": f"读取日志文件失败: {str(e)}", "success": False}

        log_with_user('info', f"返回日志记录 {len(logs)} 条", admin_user)
        logger.info(f"成功返回 {len(logs)} 条日志记录")

        return {
            "logs": logs,
            "log_file": latest_log_file,
            "total_lines": len(logs),
            "success": True
        }

    except Exception as e:
        logger.error(f"获取系统日志失败: {str(e)}")
        log_with_user('error', f"获取系统日志失败: {str(e)}", admin_user)
        return {"logs": [], "message": f"获取系统日志失败: {str(e)}", "success": False}

@router.get('/admin/stats')
def get_system_stats(admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取系统统计信息（管理员专用）"""
    from app.repositories.db_manager import db_manager
    try:
        log_with_user('info', "查询系统统计信息", admin_user)

        stats = {
            "users": {
                "total": 0,
                "active_today": 0
            },
            "cookies": {
                "total": 0,
                "enabled": 0
            },
            "cards": {
                "total": 0,
                "enabled": 0
            },
            "system": {
                "uptime": "未知",
                "version": "1.0.0"
            }
        }

        # 用户统计
        all_users = db_manager.get_all_users()
        stats["users"]["total"] = len(all_users)

        # Cookie统计
        all_cookies = db_manager.get_all_cookies()
        stats["cookies"]["total"] = len(all_cookies)

        # 卡券统计
        all_cards = db_manager.get_all_cards()
        if all_cards:
            stats["cards"]["total"] = len(all_cards)
            stats["cards"]["enabled"] = len([card for card in all_cards if card.get('enabled', True)])

        log_with_user('info', "系统统计信息查询完成", admin_user)
        return stats

    except Exception as e:
        log_with_user('error', f"获取系统统计信息失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

# ------------------------- 指定商品回复接口 -------------------------

@router.get("/itemReplays")
def get_all_item_replies(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的所有商品回复信息"""
    try:
        # 只返回当前用户的商品信息
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        all_items = []
        for cookie_id in user_cookies.keys():
            items = db_manager.get_itemReplays_by_cookie(cookie_id)
            all_items.extend(items)

        return {"items": all_items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品回复信息失败: {str(e)}")

@router.get("/itemReplays/cookie/{cookie_id}")
def get_item_replies_by_cookie(cookie_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取指定Cookie的商品信息"""
    try:
        # 检查cookie是否属于当前用户
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager
        user_cookies = db_manager.get_all_cookies(user_id)

        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        items = db_manager.get_itemReplays_by_cookie(cookie_id)
        return {"items": items}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品信息失败: {str(e)}")

@router.put("/item-reply/{cookie_id}/{item_id}")
def update_item_reply(
    cookie_id: str,
    item_id: str,
    data: dict,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    更新指定账号和商品的回复内容
    """
    try:
        user_id = current_user['user_id']
        from app.repositories.db_manager import db_manager

        # 验证cookie是否属于用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        reply_content = data.get("reply_content", "").strip()
        if not reply_content:
            raise HTTPException(status_code=400, detail="回复内容不能为空")

        db_manager.update_item_reply(cookie_id=cookie_id, item_id=item_id, reply_content=reply_content)

        return {"message": "商品回复更新成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"更新商品回复失败: {str(e)}")

@router.delete("/item-reply/{cookie_id}/{item_id}")
def delete_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    删除指定账号cookie_id和商品item_id的商品回复
    """
    try:
        user_id = current_user['user_id']
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        success = db_manager.delete_item_reply(cookie_id, item_id)
        if not success:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return {"message": "商品回复删除成功"}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"删除商品回复失败: {str(e)}")

class ItemReplyToDelete(BaseModel):
    cookie_id: str
    item_id: str


class ItemReplyBatchDeleteRequest(BaseModel):
    items: List[ItemReplyToDelete]

@router.delete("/item-reply/batch")
async def batch_delete_item_reply(
    req: ItemReplyBatchDeleteRequest,
    current_user: Dict[str, Any] = Depends(get_current_user)
):
    """
    批量删除商品回复
    """
    user_id = current_user['user_id']
    from app.repositories.db_manager import db_manager

    # 先校验当前用户是否有权限删除每个cookie对应的回复
    user_cookies = db_manager.get_all_cookies(user_id)
    for item in req.items:
        if item.cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail=f"无权限访问Cookie {item.cookie_id}")

    result = db_manager.batch_delete_item_replies([item.dict() for item in req.items])
    return {
        "success_count": result["success_count"],
        "failed_count": result["failed_count"]
    }

@router.get("/item-reply/{cookie_id}/{item_id}")
def get_item_reply(cookie_id: str, item_id: str, current_user: Dict[str, Any] = Depends(get_current_user)):
    """
    获取指定账号cookie_id和商品item_id的商品回复内容
    """
    try:
        user_id = current_user['user_id']
        # 校验cookie_id是否属于当前用户
        user_cookies = db_manager.get_all_cookies(user_id)
        if cookie_id not in user_cookies:
            raise HTTPException(status_code=403, detail="无权限访问该Cookie")

        # 获取指定商品回复
        item_replies = db_manager.get_itemReplays_by_cookie(cookie_id)
        # 找对应item_id的回复
        item_reply = next((r for r in item_replies if r['item_id'] == item_id), None)

        if item_reply is None:
            raise HTTPException(status_code=404, detail="商品回复不存在")

        return item_reply

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取商品回复失败: {str(e)}")


# ------------------------- 数据库备份和恢复接口 -------------------------

@router.get('/admin/backup/download')
def download_database_backup(admin_user: Dict[str, Any] = Depends(require_admin)):
    """下载数据库备份文件（管理员专用）"""
    try:
        log_with_user('info', "请求下载数据库备份", admin_user)

        # 使用db_manager的实际数据库路径
        from app.repositories.db_manager import db_manager
        db_file_path = db_manager.db_path

        # 检查数据库文件是否存在
        if not os.path.exists(db_file_path):
            log_with_user('error', f"数据库文件不存在: {db_file_path}", admin_user)
            raise HTTPException(status_code=404, detail="数据库文件不存在")

        # 生成带时间戳的文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        download_filename = f"xianyu_backup_{timestamp}.db"

        log_with_user('info', f"开始下载数据库备份: {download_filename}", admin_user)

        return FileResponse(
            path=db_file_path,
            filename=download_filename,
            media_type='application/octet-stream'
        )

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"下载数据库备份失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@router.post('/admin/backup/upload')
async def upload_database_backup(admin_user: Dict[str, Any] = Depends(require_admin),
                                backup_file: UploadFile = File(...)):
    """上传并恢复数据库备份文件（管理员专用）"""
    try:
        log_with_user('info', f"开始上传数据库备份: {backup_file.filename}", admin_user)

        # 验证文件类型
        if not backup_file.filename.endswith('.db'):
            log_with_user('warning', f"无效的备份文件类型: {backup_file.filename}", admin_user)
            raise HTTPException(status_code=400, detail="只支持.db格式的数据库文件")

        # 验证文件大小（限制100MB）
        content = await backup_file.read()
        if len(content) > 100 * 1024 * 1024:  # 100MB
            log_with_user('warning', f"备份文件过大: {len(content)} bytes", admin_user)
            raise HTTPException(status_code=400, detail="备份文件大小不能超过100MB")

        # 验证是否为有效的SQLite数据库文件
        temp_file_path = f"temp_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"

        try:
            # 保存临时文件
            with open(temp_file_path, 'wb') as temp_file:
                temp_file.write(content)

            # 验证数据库文件完整性
            conn = sqlite3.connect(temp_file_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
            tables = cursor.fetchall()
            conn.close()

            # 检查是否包含必要的表
            table_names = [table[0] for table in tables]
            required_tables = ['users', 'cookies']  # 最基本的表

            missing_tables = [table for table in required_tables if table not in table_names]
            if missing_tables:
                log_with_user('warning', f"备份文件缺少必要的表: {missing_tables}", admin_user)
                raise HTTPException(status_code=400, detail=f"备份文件不完整，缺少表: {', '.join(missing_tables)}")

            log_with_user('info', f"备份文件验证通过，包含 {len(table_names)} 个表", admin_user)

        except sqlite3.Error as e:
            log_with_user('error', f"备份文件验证失败: {str(e)}", admin_user)
            if os.path.exists(temp_file_path):
                os.remove(temp_file_path)
            raise HTTPException(status_code=400, detail="无效的数据库文件")

        # 备份当前数据库
        from app.repositories.db_manager import db_manager
        current_db_path = db_manager.db_path

        # 生成备份文件路径（与原数据库在同一目录）
        db_dir = os.path.dirname(current_db_path)
        backup_filename = f"xianyu_data_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        backup_current_path = os.path.join(db_dir, backup_filename)

        if os.path.exists(current_db_path):
            shutil.copy2(current_db_path, backup_current_path)
            log_with_user('info', f"当前数据库已备份为: {backup_current_path}", admin_user)

        # 关闭当前数据库连接
        if hasattr(db_manager, 'conn') and db_manager.conn:
            db_manager.conn.close()
            log_with_user('info', "已关闭当前数据库连接", admin_user)

        # 替换数据库文件
        shutil.move(temp_file_path, current_db_path)
        log_with_user('info', f"数据库文件已替换: {current_db_path}", admin_user)

        # 重新初始化数据库连接（使用原有的db_path）
        db_manager.__init__(db_manager.db_path)
        log_with_user('info', "数据库连接已重新初始化", admin_user)

        # 验证新数据库
        try:
            test_users = db_manager.get_all_users()
            log_with_user('info', f"数据库恢复成功，包含 {len(test_users)} 个用户", admin_user)
        except Exception as e:
            log_with_user('error', f"数据库恢复后验证失败: {str(e)}", admin_user)
            # 如果验证失败，尝试恢复原数据库
            if os.path.exists(backup_current_path):
                shutil.copy2(backup_current_path, current_db_path)
                db_manager.__init__()
                log_with_user('info', "已恢复原数据库", admin_user)
            raise HTTPException(status_code=500, detail="数据库恢复失败，已回滚到原数据库")

        return {
            "success": True,
            "message": "数据库恢复成功",
            "backup_file": backup_current_path,
            "user_count": len(test_users)
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"上传数据库备份失败: {str(e)}", admin_user)
        # 清理临时文件
        if 'temp_file_path' in locals() and os.path.exists(temp_file_path):
            os.remove(temp_file_path)
        raise HTTPException(status_code=500, detail=str(e))

@router.get('/admin/backup/list')
def list_backup_files(admin_user: Dict[str, Any] = Depends(require_admin)):
    """列出服务器上的备份文件（管理员专用）"""
    try:
        log_with_user('info', "查询备份文件列表", admin_user)

        # 查找备份文件（在data目录中）
        backup_files = glob.glob("data/xianyu_data_backup_*.db")

        backup_list = []
        for file_path in backup_files:
            try:
                stat = os.stat(file_path)
                backup_list.append({
                    'filename': os.path.basename(file_path),
                    'size': stat.st_size,
                    'size_mb': round(stat.st_size / (1024 * 1024), 2),
                    'created_time': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                    'modified_time': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })
            except Exception as e:
                log_with_user('warning', f"读取备份文件信息失败: {file_path} - {str(e)}", admin_user)

        # 按修改时间倒序排列
        backup_list.sort(key=lambda x: x['modified_time'], reverse=True)

        log_with_user('info', f"找到 {len(backup_list)} 个备份文件", admin_user)

        return {
            "backups": backup_list,
            "total": len(backup_list)
        }

    except Exception as e:
        log_with_user('error', f"查询备份文件列表失败: {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# ------------------------- 数据管理接口 -------------------------

@router.get('/admin/data/{table_name}')
def get_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """获取指定表的所有数据（管理员专用）"""
    from app.repositories.db_manager import db_manager
    try:
        log_with_user('info', f"查询表数据: {table_name}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', "item_replay"
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试访问不允许的表: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许访问该表")

        # 获取表数据
        data, columns = db_manager.get_table_data(table_name)

        log_with_user('info', f"表 {table_name} 查询成功，共 {len(data)} 条记录", admin_user)

        return {
            "success": True,
            "data": data,
            "columns": columns,
            "count": len(data)
        }

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"查询表数据失败: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@router.delete('/admin/data/{table_name}/{record_id}')
def delete_table_record(table_name: str, record_id: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """删除指定表的指定记录（管理员专用）"""
    from app.repositories.db_manager import db_manager
    try:
        log_with_user('info', f"删除表记录: {table_name}.{record_id}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'users', 'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders','item_replay'
        ]

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试删除不允许的表记录: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许操作该表")

        # 特殊保护：不能删除管理员用户
        if table_name == 'users' and record_id == str(admin_user['user_id']):
            log_with_user('warning', "尝试删除管理员自己", admin_user)
            raise HTTPException(status_code=400, detail="不能删除管理员自己")

        # 删除记录
        success = db_manager.delete_table_record(table_name, record_id)

        if success:
            log_with_user('info', f"表记录删除成功: {table_name}.{record_id}", admin_user)
            return {"success": True, "message": "删除成功"}
        else:
            log_with_user('warning', f"表记录删除失败: {table_name}.{record_id}", admin_user)
            raise HTTPException(status_code=400, detail="删除失败，记录可能不存在")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"删除表记录异常: {table_name}.{record_id} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))

@router.delete('/admin/data/{table_name}')
def clear_table_data(table_name: str, admin_user: Dict[str, Any] = Depends(require_admin)):
    """清空指定表的所有数据（管理员专用）"""
    from app.repositories.db_manager import db_manager
    try:
        log_with_user('info', f"清空表数据: {table_name}", admin_user)

        # 验证表名安全性
        allowed_tables = [
            'cookies', 'cookie_status', 'keywords', 'default_replies', 'default_reply_records',
            'ai_reply_settings', 'ai_conversations', 'ai_item_cache', 'item_info',
            'message_notifications', 'cards', 'delivery_rules', 'notification_channels',
            'user_settings', 'system_settings', 'email_verifications', 'captcha_codes', 'orders', 'item_replay',
            'risk_control_logs'
        ]

        # 不允许清空用户表
        if table_name == 'users':
            log_with_user('warning', "尝试清空用户表", admin_user)
            raise HTTPException(status_code=400, detail="不允许清空用户表")

        if table_name not in allowed_tables:
            log_with_user('warning', f"尝试清空不允许的表: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="不允许清空该表")

        # 清空表数据
        success = db_manager.clear_table_data(table_name)

        if success:
            log_with_user('info', f"表数据清空成功: {table_name}", admin_user)
            return {"success": True, "message": "清空成功"}
        else:
            log_with_user('warning', f"表数据清空失败: {table_name}", admin_user)
            raise HTTPException(status_code=400, detail="清空失败")

    except HTTPException:
        raise
    except Exception as e:
        log_with_user('error', f"清空表数据异常: {table_name} - {str(e)}", admin_user)
        raise HTTPException(status_code=500, detail=str(e))


# 商品多规格管理API
@router.put("/items/{cookie_id}/{item_id}/multi-spec")
def update_item_multi_spec(cookie_id: str, item_id: str, spec_data: dict, _: None = Depends(require_auth)):
    """更新商品的多规格状态"""
    try:
        from app.repositories.db_manager import db_manager

        is_multi_spec = spec_data.get('is_multi_spec', False)

        success = db_manager.update_item_multi_spec_status(cookie_id, item_id, is_multi_spec)

        if success:
            return {"message": f"商品多规格状态已{'开启' if is_multi_spec else '关闭'}"}
        else:
            raise HTTPException(status_code=404, detail="商品不存在")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# 商品多数量发货管理API
@router.put("/items/{cookie_id}/{item_id}/multi-quantity-delivery")
def update_item_multi_quantity_delivery(cookie_id: str, item_id: str, delivery_data: dict, _: None = Depends(require_auth)):
    """更新商品的多数量发货状态"""
    try:
        from app.repositories.db_manager import db_manager

        multi_quantity_delivery = delivery_data.get('multi_quantity_delivery', False)

        success = db_manager.update_item_multi_quantity_delivery_status(cookie_id, item_id, multi_quantity_delivery)

        if success:
            return {"message": f"商品多数量发货状态已{'开启' if multi_quantity_delivery else '关闭'}"}
        else:
            raise HTTPException(status_code=404, detail="商品不存在")

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))





# ==================== 订单管理接口 ====================

@router.get('/api/orders')
def get_user_orders(current_user: Dict[str, Any] = Depends(get_current_user)):
    """获取当前用户的订单信息"""
    try:
        from app.repositories.db_manager import db_manager

        user_id = current_user['user_id']
        log_with_user('info', "查询用户订单信息", current_user)

        # 获取用户的所有Cookie
        user_cookies = db_manager.get_all_cookies(user_id)

        # 获取所有订单数据
        all_orders = []
        for cookie_id in user_cookies.keys():
            orders = db_manager.get_orders_by_cookie(cookie_id, limit=1000)  # 增加限制数量
            # 为每个订单添加cookie_id信息
            for order in orders:
                order['cookie_id'] = cookie_id
                all_orders.append(order)

        # 按创建时间倒序排列
        all_orders.sort(key=lambda x: x.get('created_at', ''), reverse=True)

        log_with_user('info', f"用户订单查询成功，共 {len(all_orders)} 条记录", current_user)
        return {"success": True, "data": all_orders}

    except Exception as e:
        log_with_user('error', f"查询用户订单失败: {str(e)}", current_user)
        raise HTTPException(status_code=500, detail=f"查询订单失败: {str(e)}")


# 移除自动启动，由Start.py或手动启动
# if __name__ == "__main__":
#     uvicorn.run(app, host="0.0.0.0", port=8080)